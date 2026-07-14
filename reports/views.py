import csv
import io
from collections import defaultdict
from datetime import date, timedelta

from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Count, Q, Avg, OuterRef, Subquery, FloatField, ExpressionWrapper
from django.db.models.functions import ExtractYear, ExtractMonth
from django.utils import timezone
import datetime

from io import BytesIO

from accounts.decorators import admin_required
from audit_logs.services import log_view, log_export
from consultations.models import Consultation, Prescription, PrescriptionItem
from inventory.models import Medicine, StockMovement
from patients.models import Patient
from colleges.models import College
from feedback.models import ConsultationFeedback

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable,
)


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@login_required
@admin_required
def dashboard(request):
    log_view(
        user=request.user,
        module='Reports',
        description='Viewed report dashboard',
        request=request,
    )
    today = timezone.now().date()

    total_consultations     = Consultation.objects.count()
    consultations_today     = Consultation.objects.filter(created_at__gte=timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))).count()
    total_patients_active   = Patient.objects.filter(is_active=True, has_logged_in=True).count()
    total_patients_all      = Patient.objects.filter(is_active=True).count()
    total_patients_pending  = total_patients_all - total_patients_active

    top_medicines = (
        StockMovement.objects
        .filter(movement_type=StockMovement.MovementType.OUT)
        .values('medicine__name', 'medicine__unit')
        .annotate(total_dispensed=Sum('quantity'))
        .order_by('-total_dispensed')[:5]
    )

    low_stock = Medicine.objects.filter(
        quantity__lte=F('low_stock_threshold')
    ).order_by('quantity')

    return render(request, 'reports/report_dashboard.html', {
        'total_consultations':    total_consultations,
        'consultations_today':    consultations_today,
        'total_patients':         total_patients_active,
        'total_patients_pending': total_patients_pending,
        'top_medicines':          top_medicines,
        'low_stock':              low_stock,
    })


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _parse_date(value):
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def _make_aware_dt(d, hour=0, minute=0, second=0):
    """Create timezone-aware datetime from a date."""
    return timezone.make_aware(datetime.datetime.combine(d, datetime.time(hour, minute, second)))


def _build_diagnosis_queryset(keyword, date_from, date_to, patient_type, college_id):
    qs = (
        Consultation.objects
        .filter(status=Consultation.Status.COMPLETED)
                .select_related('patient', 'patient__college')
        .prefetch_related('prescriptions')
    )

    if keyword:
        qs = qs.filter(prescriptions__diagnosis__icontains=keyword).distinct()
    if date_from:
        qs = qs.filter(created_at__gte=_make_aware_dt(date_from))
    if date_to:
        qs = qs.filter(created_at__lte=_make_aware_dt(date_to, 23, 59, 59))

    if patient_type == 'student':
        qs = qs.filter(patient__college__isnull=False)
    elif patient_type == 'staff':
        qs = qs.filter(patient__college__isnull=True, patient__department__gt='')
    elif patient_type == 'instructor':
        qs = qs.filter(patient__college__isnull=True, patient__position__gt='')

    if college_id:
        qs = qs.filter(patient__college_id=college_id)

    return qs.order_by('-created_at')


# ─── DIAGNOSIS ANALYTICS ──────────────────────────────────────────────────────

@login_required
@admin_required
def diagnosis_analytics(request):
    colleges = College.objects.all().order_by('name')
    user_name = request.user.get_full_name() or request.user.username

    keyword       = request.GET.get('keyword', '').strip()
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    patient_type  = request.GET.get('patient_type', 'all')
    college_id    = request.GET.get('college_id', '').strip()
    sex           = request.GET.get('sex', '').strip()
    year_level    = request.GET.get('year_level', '').strip()

    date_from = _parse_date(date_from_str)
    date_to   = _parse_date(date_to_str)

    # ── Build base querysets for analytics (Sections 1-3) ──
    # Section 3 uses any consultation status (clinic utilization)
    base_any = Consultation.objects.all()
    # Sections 1-2 use only completed consultations (diagnosis-based)
    base_completed = Consultation.objects.filter(status=Consultation.Status.COMPLETED)

    if date_from:
        dt = _make_aware_dt(date_from)
        base_completed = base_completed.filter(created_at__gte=dt)
        base_any = base_any.filter(created_at__gte=dt)
    if date_to:
        dt = _make_aware_dt(date_to, 23, 59, 59)
        base_completed = base_completed.filter(created_at__lte=dt)
        base_any = base_any.filter(created_at__lte=dt)
    if college_id:
        base_completed = base_completed.filter(patient__college_id=college_id)
        base_any = base_any.filter(patient__college_id=college_id)

    # ── Section 1: Top Diagnoses ──
    top_diagnoses = list(
        Prescription.objects
        .filter(consultation__in=base_completed)
        .values('diagnosis')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # ── Section 2: Diagnosis Distribution by College (matrix format) ──
    # Build a college × diagnosis cross-tabulation matrix
    _diag_all_rows = (
        Prescription.objects
        .filter(
            consultation__in=base_completed,
            consultation__patient__college__isnull=False,
        )
        .values(
            'consultation__patient__college__abbreviation',
            'diagnosis',
        )
        .annotate(count=Count('id'))
        .order_by('consultation__patient__college__abbreviation', '-count')
    )

    # Top 10 diagnosis names = row headers
    diag_short_names = [d['diagnosis'][:30] for d in top_diagnoses]

    # College names (ordered) = column headers (all colleges, even with zero counts)
    diag_college_names = list(
        College.objects
        .distinct()
        .values_list('abbreviation', flat=True)
    )

    # Build matrix: {college: {diagnosis: count}}, then transpose
    _matrix = defaultdict(lambda: defaultdict(int))
    for row in _diag_all_rows:
        c_name = row['consultation__patient__college__abbreviation']
        diag = row['diagnosis'][:30]
        _matrix[c_name][diag] = row['count']

    # Transpose: make diagnoses the rows, colleges the columns
    # Include all diagnoses even if counts are zero across all colleges
    diag_matrix_rows = []
    for diag_name in diag_short_names:
        col_data = [_matrix.get(c_name, {}).get(diag_name, 0) for c_name in diag_college_names]
        diag_matrix_rows.append({
            'diagnosis': diag_name,
            'col_data': col_data,
        })

    # Also build the flat list for backward compatibility
    diagnosis_by_college = []
    for row in diag_matrix_rows:
        for i, cnt in enumerate(row['col_data']):
            if cnt:
                diagnosis_by_college.append({
                    'college': diag_college_names[i],
                    'diagnosis': row['diagnosis'],
                    'count': cnt,
                })

    # ── Section 3: Number of Patients by College (unique patients, any status) ──
    # Include all colleges, even those with zero patients
    _all_college_abbrs = list(College.objects.values_list('abbreviation', flat=True))
    _patient_counts_qs = (
        base_any
        .filter(patient__college__isnull=False)
        .values('patient__college__abbreviation')
        .annotate(count=Count('patient', distinct=True))
    )
    _counts_dict = {r['patient__college__abbreviation']: r['count'] for r in _patient_counts_qs}
    patients_by_college = sorted(
        [{'patient__college__abbreviation': abbr, 'count': _counts_dict.get(abbr, 0)}
         for abbr in _all_college_abbrs],
        key=lambda x: x['count'],
        reverse=True,
    )

    # ── Section 4: Build search queryset (diagnosis-based, completed only) ──
    consultations = _build_diagnosis_queryset(
        keyword, date_from, date_to, patient_type, college_id or None,
    )

    if sex:
        consultations = consultations.filter(patient__sex=sex)

    if year_level:
        consultations = consultations.filter(patient__profile__year_level=year_level)

    # ── Handle exports ──
    export_fmt = request.GET.get('export', '')
    export_params = _clean_export_params(request.GET.urlencode())
    if export_fmt == 'csv':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported diagnosis analytics as CSV{" — " + keyword if keyword else ""}',
            request=request,
        )
        return _diagnosis_analytics_csv(consultations)
    if export_fmt == 'pdf':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported diagnosis analytics as PDF{" — " + keyword if keyword else ""}',
            request=request,
        )
        return _diagnosis_analytics_pdf(consultations, keyword, date_from, date_to,
                                        patient_type, college_id, user_name)
    if export_fmt == 'excel':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported diagnosis analytics as Excel{" — " + keyword if keyword else ""}',
            request=request,
        )
        return _diagnosis_analytics_excel(consultations)

    # ── Compute stats for Section 4 ──
    total_affected = consultations.values('patient').distinct().count()
    consultations_count = consultations.count()

    by_type = {
        'student':    consultations.filter(patient__college__isnull=False)
                                   .values('patient').distinct().count(),
        'staff':      consultations.filter(patient__college__isnull=True,
                                           patient__department__gt='')
                                   .values('patient').distinct().count(),
        'instructor': consultations.filter(patient__college__isnull=True,
                                           patient__position__gt='')
                                   .values('patient').distinct().count(),
    }
    by_type['other'] = max(0, total_affected - sum(by_type.values()))

    by_college = (
        consultations
        .filter(patient__college__isnull=False)
        .values('patient__college__abbreviation', 'patient__college__name')
        .annotate(count=Count('patient', distinct=True))
        .order_by('-count')
    )

    return render(request, 'reports/diagnosis_analytics.html', {
        'consultations':         consultations,
        'consultations_count':   consultations_count,
        'colleges':              colleges,
        'keyword':               keyword,
        'date_from':             date_from_str,
        'date_to':               date_to_str,
        'patient_type':          patient_type,
        'college_id':            college_id,
        'sex':                   sex,
        'year_level':            year_level,
        'total_affected':        total_affected,
        'by_type':               by_type,
        'by_college':            by_college,
        'export_params':         export_params,
        # Analytics sections
        'top_diagnoses':         top_diagnoses,
        'diagnosis_by_college':  diagnosis_by_college,
        'diag_col_names':        diag_college_names,
        'diag_matrix_rows':      diag_matrix_rows,
        'patients_by_college':   patients_by_college,
        'has_filters':           any([keyword, date_from_str, date_to_str,
                                      patient_type != 'all', college_id,
                                      sex, year_level]),
    })


# ─── DIAGNOSIS FULL REPORT ────────────────────────────────────────────────────

@login_required
@admin_required
def diagnosis_full_report(request):
    """Full diagnosis report page with all diagnoses, filters, and PDF export."""
    colleges = College.objects.all().order_by('name')
    user_name = request.user.get_full_name() or request.user.username

    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    college_id    = request.GET.get('college_id', '').strip()
    section       = request.GET.get('section', 'all').strip()

    date_from = _parse_date(date_from_str)
    date_to   = _parse_date(date_to_str)

    has_filters = bool(date_from_str or date_to_str or college_id)

    # ── Build base querysets ──
    base_completed = Consultation.objects.filter(status=Consultation.Status.COMPLETED)

    if date_from:
        dt = _make_aware_dt(date_from)
        base_completed = base_completed.filter(created_at__gte=dt)
    if date_to:
        dt = _make_aware_dt(date_to, 23, 59, 59)
        base_completed = base_completed.filter(created_at__lte=dt)
    if college_id:
        base_completed = base_completed.filter(patient__college_id=college_id)

    # ── ALL Diagnoses (ranked, no limit) ──
    all_diagnoses = list(
        Prescription.objects
        .filter(consultation__in=base_completed)
        .values('diagnosis')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # ── Diagnosis Distribution by College (full matrix with ALL diagnoses) ──
    _diag_all_rows = (
        Prescription.objects
        .filter(
            consultation__in=base_completed,
            consultation__patient__college__isnull=False,
        )
        .values(
            'consultation__patient__college__abbreviation',
            'diagnosis',
        )
        .annotate(count=Count('id'))
        .order_by('consultation__patient__college__abbreviation', '-count')
    )

    # All diagnosis names = row headers
    full_diag_names = [d['diagnosis'][:40] for d in all_diagnoses]

    # College names (ordered) = column headers (all colleges, even with zero counts)
    full_college_names = list(
        College.objects
        .distinct()
        .values_list('abbreviation', flat=True)
    )

    # Build matrix: {college: {diagnosis: count}}, then transpose
    _matrix = defaultdict(lambda: defaultdict(int))
    for row in _diag_all_rows:
        c_name = row['consultation__patient__college__abbreviation']
        diag = row['diagnosis'][:40]
        _matrix[c_name][diag] = row['count']

    # Transpose: make diagnoses the rows, colleges the columns
    # Include all diagnoses even if counts are zero across all colleges
    full_matrix_rows = []
    for diag_name in full_diag_names:
        col_data = [_matrix.get(c_name, {}).get(diag_name, 0) for c_name in full_college_names]
        full_matrix_rows.append({
            'diagnosis': diag_name,
            'col_data': col_data,
        })

    # ── Patients by College (any status) ──
    base_any = Consultation.objects.all()
    if date_from:
        base_any = base_any.filter(created_at__gte=_make_aware_dt(date_from))
    if date_to:
        base_any = base_any.filter(created_at__lte=_make_aware_dt(date_to, 23, 59, 59))
    if college_id:
        base_any = base_any.filter(patient__college_id=college_id)

    # Include all colleges, even those with zero patients
    _all_college_abbrs = list(College.objects.values_list('abbreviation', flat=True))
    _patient_counts_qs = (
        base_any
        .filter(patient__college__isnull=False)
        .values('patient__college__abbreviation')
        .annotate(count=Count('patient', distinct=True))
    )
    _counts_dict = {r['patient__college__abbreviation']: r['count'] for r in _patient_counts_qs}
    patients_by_college = sorted(
        [{'patient__college__abbreviation': abbr, 'count': _counts_dict.get(abbr, 0)}
         for abbr in _all_college_abbrs],
        key=lambda x: x['count'],
        reverse=True,
    )

    # ── Summary stats ──
    total_all_diagnoses = len(all_diagnoses)
    total_diagnosis_count = sum(d['count'] for d in all_diagnoses)

    # ── Handle export ──
    export_fmt = request.GET.get('export', '')
    if export_fmt == 'pdf':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported full diagnosis report as PDF (section={section})',
            request=request,
        )
        if section == 'diagnoses':
            return _diagnoses_pdf(
                all_diagnoses, date_from, date_to,
                college_id, user_name,
            )
        elif section == 'matrix':
            return _matrix_pdf(
                full_matrix_rows, full_college_names, date_from, date_to,
                college_id, user_name,
            )
        elif section == 'patients':
            return _patients_pdf(
                patients_by_college, date_from, date_to,
                college_id, user_name,
            )
        else:
            return _diagnosis_full_report_pdf(
                all_diagnoses, full_matrix_rows, full_college_names,
                patients_by_college, date_from, date_to,
                college_id, user_name,
            )

    return render(request, 'reports/diagnosis_full_report.html', {
        'colleges':            colleges,
        'date_from':           date_from_str,
        'date_to':             date_to_str,
        'college_id':          college_id,
        'section':             section,
        'has_filters':         has_filters,
        'all_diagnoses':       all_diagnoses,
        'total_all_diagnoses': total_all_diagnoses,
        'total_diagnosis_count': total_diagnosis_count,
        'full_col_names':      full_college_names,
        'full_matrix_rows':    full_matrix_rows,
        'patients_by_college': patients_by_college,
    })


def _diagnosis_full_report_pdf(all_diagnoses, full_matrix_rows, full_college_names,
                                patients_by_college, date_from, date_to,
                                college_id, user_name=None):
    """Generate a professional PDF for the full diagnosis report."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    # ── Structured metadata ──
    meta = [f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}']
    if date_from or date_to:
        period_from = date_from.strftime('%B %d, %Y') if date_from else '—'
        period_to = date_to.strftime('%B %d, %Y') if date_to else '—'
        meta.append(f'<b>Report Period:</b>  {period_from} &mdash; {period_to}')
    if college_id:
        college = College.objects.filter(pk=college_id).first()
        if college:
            meta.append(f'<b>College:</b>  {college.abbreviation} &mdash; {college.name}')

    story = []
    story.extend(_pdf_header_block('Full Diagnosis Report', meta))

    # ── All Diagnoses (ranked) ──
    story.append(Paragraph('All Diagnoses (Ranked)', s['section_title']))
    story.append(_pdf_make_table(
        ['Rank', 'Diagnosis', 'Cases'],
        [[str(i + 1), d['diagnosis'][:90], str(d['count'])]
         for i, d in enumerate(all_diagnoses)],
        col_widths=[1.5*cm, 11*cm, 4.5*cm],
        aligns=['center', 'left', 'right'],
    ))
    story.append(Spacer(1, 8))

    # ── Diagnosis Distribution by College (transposed: diagnoses as rows) ──
    if full_matrix_rows:
        story.append(Paragraph('Diagnosis Distribution by College', s['section_title']))
        # Build table with diagnosis + all college columns
        n = len(full_college_names)
        matrix_headers = ['Diagnosis'] + [name[:16] for name in full_college_names]
        matrix_rows = []
        for row in full_matrix_rows:
            matrix_rows.append(
                [row['diagnosis'][:60]] + [str(v) if v else '—' for v in row['col_data']]
            )
        # Auto-adjust: proportional column widths to fit the page
        # Diagnosis column gets ~35% (up to 5cm), college count columns share the rest
        page_w = A4[0] - 4 * cm  # usable width ≈ 17cm
        diag_w = min(5 * cm, page_w * 0.35)
        count_w = (page_w - diag_w) / max(n, 1)
        story.append(_pdf_make_table(
            matrix_headers,
            matrix_rows,
            col_widths=[diag_w] + [count_w] * n,
            aligns=['left'] + ['right'] * n,
        ))
        story.append(Spacer(1, 8))

    # ── Patients by College ──
    if patients_by_college:
        story.append(Paragraph('Patients by College', s['section_title']))
        story.append(_pdf_make_table(
            ['College', 'Patients'],
            [[r['patient__college__abbreviation'], str(r['count'])]
             for r in patients_by_college],
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
        ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'full_diagnosis_report_{date.today()}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ─── SECTION-SPECIFIC PDF EXPORTS ────────────────────────────────────────────

def _diagnoses_pdf(all_diagnoses, date_from, date_to,
                   college_id, user_name=None):
    """PDF with only the ranked diagnoses list."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    meta = [f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}']
    if date_from or date_to:
        period_from = date_from.strftime('%B %d, %Y') if date_from else '—'
        period_to = date_to.strftime('%B %d, %Y') if date_to else '—'
        meta.append(f'<b>Report Period:</b>  {period_from} &mdash; {period_to}')
    if college_id:
        college = College.objects.filter(pk=college_id).first()
        if college:
            meta.append(f'<b>College:</b>  {college.abbreviation} &mdash; {college.name}')

    story = []
    story.extend(_pdf_header_block('Diagnoses (Ranked)', meta))

    story.append(Paragraph('All Diagnoses (Ranked)', s['section_title']))
    story.append(_pdf_make_table(
        ['Rank', 'Diagnosis', 'Cases'],
        [[str(i + 1), d['diagnosis'][:90], str(d['count'])]
         for i, d in enumerate(all_diagnoses)],
        col_widths=[1.5*cm, 11*cm, 4.5*cm],
        aligns=['center', 'left', 'right'],
    ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'diagnoses_{date.today()}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _matrix_pdf(full_matrix_rows, full_college_names, date_from, date_to,
                college_id, user_name=None):
    """PDF with only the Diagnosis Distribution by College matrix."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    meta = [f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}']
    if date_from or date_to:
        period_from = date_from.strftime('%B %d, %Y') if date_from else '—'
        period_to = date_to.strftime('%B %d, %Y') if date_to else '—'
        meta.append(f'<b>Report Period:</b>  {period_from} &mdash; {period_to}')
    if college_id:
        college = College.objects.filter(pk=college_id).first()
        if college:
            meta.append(f'<b>College:</b>  {college.abbreviation} &mdash; {college.name}')

    story = []
    story.extend(_pdf_header_block('Diagnosis Distribution by College', meta))

    if full_matrix_rows:
        story.append(Paragraph('Diagnosis Distribution by College', s['section_title']))
        n = len(full_college_names)
        matrix_headers = ['Diagnosis'] + [name[:16] for name in full_college_names]
        matrix_rows = []
        for row in full_matrix_rows:
            matrix_rows.append(
                [row['diagnosis'][:60]] + [str(v) if v else '—' for v in row['col_data']]
            )
        page_w = A4[0] - 4 * cm
        diag_w = min(5 * cm, page_w * 0.35)
        count_w = (page_w - diag_w) / max(n, 1)
        story.append(_pdf_make_table(
            matrix_headers, matrix_rows,
            col_widths=[diag_w] + [count_w] * n,
            aligns=['left'] + ['right'] * n,
        ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'diagnosis_by_college_{date.today()}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _patients_pdf(patients_by_college, date_from, date_to,
                  college_id, user_name=None):
    """PDF with only the Patients by College table."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    meta = [f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}']
    if date_from or date_to:
        period_from = date_from.strftime('%B %d, %Y') if date_from else '—'
        period_to = date_to.strftime('%B %d, %Y') if date_to else '—'
        meta.append(f'<b>Report Period:</b>  {period_from} &mdash; {period_to}')
    if college_id:
        college = College.objects.filter(pk=college_id).first()
        if college:
            meta.append(f'<b>College:</b>  {college.abbreviation} &mdash; {college.name}')

    story = []
    story.extend(_pdf_header_block('Patients by College', meta))

    if patients_by_college:
        story.append(Paragraph('Patients by College', s['section_title']))
        story.append(_pdf_make_table(
            ['College', 'Patients'],
            [[r['patient__college__abbreviation'], str(r['count'])]
             for r in patients_by_college],
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
        ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'patients_by_college_{date.today()}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _diagnosis_analytics_pdf(consultations, keyword, date_from, date_to, patient_type, college_id, user_name=None):
    """Generate a professional PDF for the diagnosis analytics report."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    # ── Compute summary stats ──
    total_affected = consultations.values('patient').distinct().count()

    student_count = consultations.filter(
        patient__college__isnull=False
    ).values('patient').distinct().count()
    staff_count = consultations.filter(
        patient__college__isnull=True, patient__department__gt=''
    ).values('patient').distinct().count()
    instructor_count = consultations.filter(
        patient__college__isnull=True, patient__position__gt=''
    ).values('patient').distinct().count()

    total_consultations = consultations.count()

    by_college = list(
        consultations
        .filter(patient__college__isnull=False)
        .values('patient__college__abbreviation', 'patient__college__name')
        .annotate(count=Count('patient', distinct=True))
        .order_by('-count')
    )

    # ── Structured metadata ──
    meta = [f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}']
    if date_from or date_to:
        period_from = date_from.strftime('%B %d, %Y') if date_from else '—'
        period_to = date_to.strftime('%B %d, %Y') if date_to else '—'
        meta.append(f'<b>Report Period:</b>  {period_from} &mdash; {period_to}')
    if keyword:
        meta.append(f'<b>Search Keyword:</b>  {keyword}')
    if patient_type and patient_type != 'all':
        meta.append(f'<b>Patient Category:</b>  {patient_type.capitalize()}')

    story = []
    story.extend(_pdf_header_block('Disease Summary Report', meta))

    # ── Summary Overview ──
    story.append(Paragraph('Summary Overview', s['section_title']))
    story.append(_pdf_make_table(
        ['Metric', 'Count'],
        [
            ['Patients Diagnosed', str(total_affected)],
            ['Total Consultations', str(total_consultations)],
            ['Student Patients', str(student_count)],
            ['Instructor Patients', str(instructor_count)],
            ['Staff Patients', str(staff_count)],
        ],
        col_widths=[12*cm, 5*cm],
        aligns=['left', 'right'],
        h_align='CENTER',
    ))
    story.append(Spacer(1, 8))

    # ── Distribution by College ──
    if by_college:
        story.append(Paragraph('Distribution by College', s['section_title']))
        story.append(_pdf_make_table(
            ['College', 'Patients Diagnosed'],
            [[f"{r['patient__college__abbreviation']} &mdash; {r['patient__college__name']}",
              str(r['count'])]
             for r in by_college],
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
        ))

        if total_consultations > 100:
            story.append(Spacer(1, 6))
            story.append(Paragraph(
                f'Showing the first 100 of {total_consultations} records. '
                f'Apply date filters to narrow the range.',
                s['report_meta'],
            ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'diagnosis_analytics_{date.today()}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _diagnosis_analytics_csv(consultations):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="diagnosis_analytics_{date.today()}.csv"'
    )

    # Annotate each consultation row with the first prescription's diagnosis
    # and treatment_plan via subquery, avoiding N+1 queries that would occur
    # if we relied on prefetch_related (which .iterator() silently drops).
    first_rx = Prescription.objects.filter(
        consultation=OuterRef('pk'),
    ).order_by('prescribed_at').values('diagnosis', 'treatment_plan')[:1]

    consultations = consultations.annotate(
        _first_diagnosis=Subquery(first_rx.values('diagnosis')),
        _first_treatment=Subquery(first_rx.values('treatment_plan')),
    )

    writer = csv.writer(response)
    writer.writerow([
        'Consultation #', 'Date', 'Patient Name', 'Patient ID',
        'Type', 'College / Department', 'Diagnosis', 'Treatment Plan',
    ])
    for c in consultations.iterator():
        p  = c.patient
        if p.college:
            p_type, p_org = 'Student', p.college.abbreviation
        elif p.department:
            p_type, p_org = 'Staff', p.department
        else:
            p_type, p_org = 'Instructor', p.position or '—'
        writer.writerow([
            c.pk,
            c.created_at.strftime('%Y-%m-%d'),
            p.get_full_name(),
            p.patient_id,
            p_type,
            p_org,
            c._first_diagnosis or '—',
            c._first_treatment or '—',
        ])
    return response


def _diagnosis_analytics_excel(consultations):
    """Generate an Excel (.xlsx) export for diagnosis analytics search results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    # Annotate each consultation row with the first prescription's diagnosis
    # and treatment_plan via subquery, matching the CSV approach.
    first_rx = Prescription.objects.filter(
        consultation=OuterRef('pk'),
    ).order_by('prescribed_at').values('diagnosis', 'treatment_plan')[:1]

    consultations = consultations.annotate(
        _first_diagnosis=Subquery(first_rx.values('diagnosis')),
        _first_treatment=Subquery(first_rx.values('treatment_plan')),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Diagnosis Analytics'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='0078d4')

    # Title row
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='Diagnosis Analytics — Search Results').font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

    # Header row
    headers = ['#', 'Date', 'Patient Name', 'Patient ID', 'Sex',
               'Type', 'College / Department', 'Diagnosis', 'Treatment Plan']
    for col_idx, col in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, c in enumerate(consultations.iterator(), start=4):
        p = c.patient
        if p.college:
            p_type, p_org = 'Student', p.college.abbreviation
        elif p.department:
            p_type, p_org = 'Staff', p.department
        else:
            p_type, p_org = 'Instructor', p.position or '—'

        sex = {'M': 'Male', 'F': 'Female'}.get(p.sex, '—')

        ws.cell(row=row_idx, column=1, value=c.pk)
        ws.cell(row=row_idx, column=2, value=c.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=3, value=p.get_full_name())
        ws.cell(row=row_idx, column=4, value=p.patient_id)
        ws.cell(row=row_idx, column=5, value=sex)
        ws.cell(row=row_idx, column=6, value=p_type)
        ws.cell(row=row_idx, column=7, value=p_org)
        ws.cell(row=row_idx, column=8, value=c._first_diagnosis or '—')
        ws.cell(row=row_idx, column=9, value=c._first_treatment or '—')

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="diagnosis_analytics_{date.today()}.xlsx"'
    )
    return response


# ─── FEEDBACK REPORT ───────────────────────────────────────────────────────────

@login_required
@admin_required
def feedback_report(request):
    """Patient feedback summary with PDF/CSV export."""
    user_name = request.user.get_full_name() or request.user.username
    feedbacks = ConsultationFeedback.objects.select_related(
        'consultation__patient'
    ).order_by('-created_at')

    # ── Filters ──
    search = request.GET.get('search', '').strip()
    if search:
        feedbacks = feedbacks.filter(
            Q(consultation__patient__first_name__icontains=search) |
            Q(consultation__patient__last_name__icontains=search) |
            Q(consultation__patient__patient_id__icontains=search) |
            Q(comment__icontains=search)
        )

    rating = request.GET.get('rating', '')
    if rating in ('1','2','3','4','5'):
        feedbacks = feedbacks.filter(rating=int(rating))

    export_fmt = request.GET.get('export', '')
    if export_fmt == 'pdf':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported feedback report as PDF{" — " + search if search else ""}',
            request=request,
        )
        return _feedback_pdf(feedbacks, search, rating, user_name)
    if export_fmt == 'csv':
        log_export(
            user=request.user,
            module='Reports',
            description=f'Exported feedback report as CSV{" — " + search if search else ""}',
            request=request,
        )
        return _feedback_csv(feedbacks)

    # ── Summary stats ──
    total = feedbacks.count()
    avg_rating = round(feedbacks.aggregate(avg=Avg('rating'))['avg'] or 0, 1)
    rating_dist = []
    for r in range(5, 0, -1):
        cnt = feedbacks.filter(rating=r).count()
        if cnt:
            rating_dist.append({'rating': r, 'count': cnt,
                                'pct': round(cnt / total * 100, 1) if total else 0})

    return render(request, 'reports/feedback_report.html', {
        'feedbacks':    feedbacks,
        'search':       search,
        'rating_filter': rating,
        'total':        total,
        'avg_rating':   avg_rating,
        'rating_dist':  rating_dist,
        'export_params': _clean_export_params(request.GET.urlencode()),
    })


def _feedback_csv(feedbacks):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="feedback_report_{date.today()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(['#', 'Patient Name', 'Patient ID', 'Consultation',
                     'Rating', 'Review', 'Date'])
    for idx, f in enumerate(feedbacks.iterator(), 1):
        writer.writerow([
            idx,
            f.consultation.patient.get_full_name() or '—',
            f.consultation.patient.patient_id or '—',
            f'#{f.consultation.pk}',
            f.rating,
            f.comment.strip() or '—',
            f.created_at.strftime('%Y-%m-%d') if f.created_at else '—',
        ])
    return response


# ═══════════════════════════════════════════════════════════════════════
# PDF DESIGN SYSTEM — shared layout, styles, and helpers (grayscale, formal)
# ═══════════════════════════════════════════════════════════════════════

_PDF_BLACK      = colors.black
_PDF_DARK_GRAY  = colors.HexColor('#333333')
_PDF_MID_GRAY   = colors.HexColor('#666666')
_PDF_MUTED      = colors.HexColor('#888888')
_PDF_LIGHT_GRAY = colors.HexColor('#f0f0f0')
_PDF_WHITE      = colors.white
_PDF_BORDER     = colors.HexColor('#999999')


def _pdf_styles():
    """Return a dict of shared ParagraphStyle objects for PDF generation (grayscale)."""
    base = getSampleStyleSheet()
    return {
        'clinic_name': ParagraphStyle(
            'PdfClinicName', parent=base['Heading1'],
            fontName='Times-Bold', fontSize=16, spaceAfter=2,
            textColor=_PDF_BLACK, alignment=1,
        ),
        'clinic_subtitle': ParagraphStyle(
            'PdfClinicSubtitle', parent=base['Normal'],
            fontName='Times-Roman', fontSize=9,
            textColor=_PDF_DARK_GRAY, spaceAfter=10, alignment=1,
        ),
        'report_title': ParagraphStyle(
            'PdfReportTitle', parent=base['Heading2'],
            fontName='Times-Bold', fontSize=13, spaceAfter=4,
            textColor=_PDF_BLACK,
        ),
        'report_meta': ParagraphStyle(
            'PdfReportMeta', parent=base['Normal'],
            fontName='Times-Roman', fontSize=8,
            textColor=_PDF_MID_GRAY, spaceAfter=14, leading=12,
        ),
        'section_title': ParagraphStyle(
            'PdfSectionTitle', parent=base['Heading3'],
            fontName='Times-Bold', fontSize=10, spaceAfter=6,
            spaceBefore=12, textColor=_PDF_BLACK,
        ),
        'th': ParagraphStyle(
            'PdfTH', parent=base['Normal'],
            fontName='Helvetica-Bold', fontSize=8, leading=10,
            textColor=_PDF_BLACK,
        ),
        'td': ParagraphStyle(
            'PdfTD', parent=base['Normal'],
            fontName='Times-Roman', fontSize=8, leading=11,
            textColor=_PDF_BLACK,
        ),
        'td_c': ParagraphStyle(
            'PdfTDC', parent=base['Normal'],
            fontName='Times-Roman', fontSize=8, leading=11,
            textColor=_PDF_BLACK, alignment=1,
        ),
        'td_r': ParagraphStyle(
            'PdfTDR', parent=base['Normal'],
            fontName='Times-Roman', fontSize=8, leading=11,
            textColor=_PDF_BLACK, alignment=2,
        ),
    }


def _pdf_header_block(report_title, meta_lines=None):
    """Return story elements: clinic identity + report title + metadata (grayscale)."""
    s = _pdf_styles()
    els = []
    els.append(Paragraph('NEGROS ORIENTAL STATE UNIVERSITY', s['clinic_name']))
    els.append(Paragraph(
        'University Medical-Dental Clinic',
        s['clinic_subtitle'],
    ))
    els.append(Paragraph(
        'Patient Record Management System',
        ParagraphStyle('PdfSub2', parent=s['clinic_subtitle'], spaceAfter=10),
    ))
    els.append(HRFlowable(width='100%', thickness=0.5, color=_PDF_BORDER))
    els.append(Spacer(1, 10))
    els.append(Paragraph(report_title, s['report_title']))
    if meta_lines:
        for line in meta_lines:
            els.append(Paragraph(line, s['report_meta']))
    els.append(Spacer(1, 6))
    return els


def _make_pdf_footer(user_name=None):
    """Return a page footer callback with confidentiality notice, user, and page number (grayscale)."""
    def _footer(canvas, doc):
        canvas.saveState()
        margin = 2 * cm
        canvas.setStrokeColor(_PDF_BORDER)
        canvas.setLineWidth(0.3)
        canvas.line(margin, 1.5 * cm, A4[0] - margin, 1.5 * cm)
        canvas.setFont('Times-Roman', 7)
        canvas.setFillColor(_PDF_MID_GRAY)
        canvas.drawCentredString(
            A4[0] / 2, 1.1 * cm,
            'Confidential Medical Record - For authorized clinic personnel only',
        )
        if user_name:
            canvas.drawString(margin, 0.8 * cm, f'Generated by: {user_name}')
        canvas.drawRightString(A4[0] - margin, 0.8 * cm, f'Page {doc.page}')
        canvas.restoreState()
    return _footer


def _pdf_make_table(headers, rows, col_widths=None, aligns=None, h_align='LEFT'):
    """
    Create a clean bordered table with light gray header background, black text,
    and thin gray grid lines. No colored fills.

    aligns: optional list of 'left'|'center'|'right' per column.
    h_align: horizontal table alignment — 'LEFT', 'CENTER', or 'RIGHT'.
    """
    s = _pdf_styles()
    align_map = {'left': s['td'], 'center': s['td_c'], 'right': s['td_r']}

    # Header row: each header's alignment should match its column's data alignment
    header_row = []
    for i, h in enumerate(headers):
        h_style = align_map.get(aligns[i] if aligns else 'left', s['td'])
        # Header uses Helvetica-Bold but the same alignment as data
        header_style = ParagraphStyle(
            'PdfTH_custom', fontName='Helvetica-Bold', fontSize=8, leading=10,
            textColor=_PDF_BLACK, alignment=h_style.alignment,
        )
        header_row.append(Paragraph(h, header_style))
    data = [header_row]

    for row in rows:
        data_row = []
        for i, cell in enumerate(row):
            if isinstance(cell, Paragraph):
                data_row.append(cell)
            else:
                style = align_map.get(aligns[i] if aligns else 'left', s['td'])
                data_row.append(Paragraph(str(cell) if cell is not None else '&mdash;', style))
        data.append(data_row)

    t = Table(data, colWidths=col_widths, repeatRows=1, hAlign=h_align)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), _PDF_LIGHT_GRAY),
        ('TEXTCOLOR', (0, 0), (-1, -1), _PDF_BLACK),
        ('GRID', (0, 0), (-1, -1), 0.3, _PDF_BORDER),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, 0), 0.6, _PDF_BLACK),
    ]))
    return t


# ─── PDF GENERATORS ───────────────────────────────────────────────────────────

def _pdf_build_doc():
    """Create a BytesIO buffer and SimpleDocTemplate with A4 portrait + footer."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2.5*cm,
    )
    return buf, doc


def _feedback_pdf(feedbacks, search, rating, user_name=None):
    """Generate a professional PDF for the feedback report.

    Uses aggregation queries (.count(), .aggregate()) instead of
    loading all feedback objects into memory, keeping memory usage
    constant regardless of dataset size.
    """
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    # ── Compute summary stats via DB aggregations (no in-memory loading) ──
    total = feedbacks.count()
    avg_result = feedbacks.aggregate(avg=Avg('rating'))
    avg_rating = round(avg_result['avg'], 1) if avg_result['avg'] else 0

    # ── Structured metadata ──
    meta = [f'<b>Generated Date:</b>  {timezone.now():%B %d, %Y at %I:%M %p}']
    parts = []
    if search:
        parts.append(f'Search: &ldquo;{search}&rdquo;')
    if rating:
        parts.append(f'Rating: {rating} / 5')
    if parts:
        meta.append(f'<b>Filters Applied:</b>  {" &mdash; ".join(parts)}')
    meta.append(f'<b>Total Responses:</b>  {total}')
    if total:
        meta.append(f'<b>Average Rating:</b>  {avg_rating} / 5')

    story = []
    story.extend(_pdf_header_block('Patient Satisfaction Report', meta))

    if total == 0:
        story.append(Paragraph(
            'No feedback records match the current filters.', s['td']
        ))
    else:
        # ── Rating distribution summary (using filtered counts, no iteration) ──
        story.append(Paragraph('Rating Distribution', s['section_title']))
        dist_rows = []
        for r in range(5, 0, -1):
            rating_qs = feedbacks.filter(rating=r)
            cnt = rating_qs.count()
            students = rating_qs.filter(
                consultation__patient__college__isnull=False
            ).count()
            instructor_rating = rating_qs.filter(
                consultation__patient__college__isnull=True,
                consultation__patient__position__gt='',
            ).count()
            staff = rating_qs.filter(
                consultation__patient__college__isnull=True,
                consultation__patient__position='',
                consultation__patient__department__gt='',
            ).count()
            pct = round(cnt / total * 100, 1) if total else 0
            dist_rows.append([
                f'{r} Star' + ('s' if r > 1 else ''),
                str(students), str(instructor_rating), str(staff),
                str(cnt), f'{pct}%',
            ])
        story.append(_pdf_make_table(
            ['Rating', 'Students', 'Instructor', 'Staff', 'Total', '%'],
            dist_rows,
            col_widths=[3*cm, 3*cm, 3*cm, 3*cm, 3*cm, 2*cm],
            aligns=['left', 'right', 'right', 'right', 'right', 'right'],
        ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    filename = f'feedback_report_{timezone.now():%Y%m%d}.pdf'
    return HttpResponse(
        pdf, content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ─── CUSTOM REPORT BUILDER ────────────────────────────────────────────────────

ALL_METRICS = [
    ('total_consultations',   'Total Consultations'),
    ('total_patients',        'Total Unique Patients'),
    ('completion_rate',       'Completion Rate (%)'),
    ('cancellation_rate',     'Cancellation Rate (%)'),
    ('avg_per_day',           'Average Consultations / Day'),
    ('trend',                 'Consultation Trend'),
    ('top_diagnoses',         'Top Diagnoses'),
    ('top_medicines',         'Most Prescribed Medicines'),
    ('top_diagnosis_per_college', 'Top Diagnosis per College'),
    ('cases_per_college',     'Cases per College'),
    ('cases_by_sex',          'Cases by Sex'),
    ('cases_by_patient_type', 'Cases by Patient Type'),
    ('urgency_breakdown',     'Urgency Breakdown (Triage)'),
    ('medicine_dispensed',    'Medicine Dispensing Summary'),
    ('low_stock',             'Low Stock Medicines'),
    ('new_patients',          'New Patients in Period'),
    ('repeat_patients',       'Repeat vs. New Patient Ratio'),
    ('frequent_patients',     'Most Frequent Patients'),
]


@login_required
@admin_required
def report_builder(request):
    colleges = College.objects.all().order_by('name')
    user_name = request.user.get_full_name() or request.user.username

    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    college_id    = request.GET.get('college_id', '').strip()
    keyword       = request.GET.get('keyword', '').strip()
    grouping      = request.GET.get('grouping', 'date')
    metrics       = request.GET.getlist('metrics')
    export_fmt    = request.GET.get('export', '')

    date_from = _parse_date(date_from_str)
    date_to   = _parse_date(date_to_str)

    has_query  = bool(date_from_str and date_to_str)
    date_error = None
    results    = None

    if has_query:
        if not date_from:
            date_error = 'Invalid "Date From" value.'
        elif not date_to:
            date_error = 'Invalid "Date To" value.'
        elif date_from > date_to:
            date_error = '"Date From" must be before "Date To".'
        else:
            results = _build_report_results(
                date_from, date_to, college_id or None,
                keyword, grouping, metrics,
            )
            if export_fmt == 'csv':
                return _report_csv(results, date_from, date_to)
            if export_fmt == 'excel':
                return _report_excel(results, date_from, date_to)
            if export_fmt == 'pdf':
                return _report_pdf(results, date_from, date_to, user_name)

    if not metrics:
        metrics = [m[0] for m in ALL_METRICS]

    return render(request, 'reports/report_builder.html', {
        'colleges':       colleges,
        'date_from':      date_from_str,
        'date_to':        date_to_str,
        'college_id':     college_id,
        'keyword':        keyword,
        'grouping':       grouping,
        'metrics':        metrics,
        'all_metrics':    ALL_METRICS,
        'has_query':      has_query,
        'date_error':     date_error,
        'results':        results,
        'export_params':  _clean_export_params(request.GET.urlencode()),
    })


def _clean_export_params(qs):
    for fmt in ('csv', 'excel', 'pdf'):
        qs = qs.replace(f'&export={fmt}', '').replace(f'export={fmt}&', '').replace(f'export={fmt}', '')
    return qs


def _build_report_results(date_from, date_to, college_id, keyword, grouping, metrics):
    base_qs = Consultation.objects.filter(
        created_at__gte=_make_aware_dt(date_from),
        created_at__lte=_make_aware_dt(date_to, 23, 59, 59),
    )
    if college_id:
        base_qs = base_qs.filter(patient__college_id=college_id)
    if keyword:
        base_qs = base_qs.filter(prescriptions__diagnosis__icontains=keyword).distinct()

    completed_qs  = base_qs.filter(status=Consultation.Status.COMPLETED)
    cancelled_qs  = base_qs.filter(status=Consultation.Status.CANCELLED)
    total_count   = base_qs.count()

    results = {
        'date_from': date_from,
        'date_to':   date_to,
        'grouping':  grouping,
        'metrics':   metrics,
    }

    if 'total_consultations' in metrics:
        results['total_consultations'] = total_count

    if 'total_patients' in metrics:
        results['total_patients'] = base_qs.values('patient').distinct().count()

    if 'completion_rate' in metrics:
        completed_count = completed_qs.count()
        cancelled_count = cancelled_qs.count()
        closed_count = base_qs.filter(status=Consultation.Status.CLOSED).count()
        resolved = completed_count + cancelled_count + closed_count
        results['completion_rate'] = (
            round(completed_count / resolved * 100, 1) if resolved else 0
        )

    if 'cancellation_rate' in metrics:
        completed_count = completed_qs.count()
        cancelled_count = cancelled_qs.count()
        closed_count = base_qs.filter(status=Consultation.Status.CLOSED).count()
        resolved = completed_count + cancelled_count + closed_count
        results['cancellation_rate'] = (
            round(cancelled_count / resolved * 100, 1) if resolved else 0
        )

    if 'avg_per_day' in metrics:
        days = max((date_to - date_from).days + 1, 1)
        results['avg_per_day'] = round(total_count / days, 1)

    if 'top_diagnoses' in metrics:
        results['top_diagnoses'] = list(
            Prescription.objects
            .filter(consultation__in=completed_qs)
            .values('diagnosis')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

    if 'top_medicines' in metrics:
        results['top_medicines'] = list(
            PrescriptionItem.objects
            .filter(prescription__consultation__in=completed_qs)
            .exclude(medicine_name='')
            .values('medicine_name')
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

    if 'cases_per_college' in metrics:
        results['cases_per_college'] = list(
            base_qs
            .filter(patient__college__isnull=False)
            .values('patient__college__abbreviation', 'patient__college__name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

    if 'cases_by_sex' in metrics:
        results['cases_by_sex'] = list(
            base_qs.values('patient__sex').annotate(count=Count('id')).order_by('-count')
        )

    if 'cases_by_patient_type' in metrics:
        students    = base_qs.filter(patient__college__isnull=False).count()
        staff       = base_qs.filter(patient__college__isnull=True,
                                     patient__department__gt='').count()
        instructors = base_qs.filter(patient__college__isnull=True,
                                     patient__position__gt='').count()
        classified  = students + staff + instructors
        results['cases_by_patient_type'] = {
            'students':    students,
            'staff':       staff,
            'instructors': instructors,
            'other':       max(0, total_count - classified),
        }

    if 'urgency_breakdown' in metrics:
        from consultations.models import Triage
        results['urgency_breakdown'] = list(
            Triage.objects.filter(consultation__in=base_qs)
            .values('urgency')
            .annotate(count=Count('consultation', distinct=True))
            .order_by('-count')
        )

    if 'medicine_dispensed' in metrics:
        results['medicine_dispensed'] = list(
            StockMovement.objects
            .filter(
                movement_type=StockMovement.MovementType.OUT,
                created_at__gte=_make_aware_dt(date_from),
                created_at__lte=_make_aware_dt(date_to, 23, 59, 59),
            )
            .values('medicine__name', 'medicine__unit')
            .annotate(total_dispensed=Sum('quantity'))
            .order_by('-total_dispensed')[:15]
        )

    if 'new_patients' in metrics:
        results['new_patients'] = Patient.objects.filter(
            is_active=True,
            created_at__gte=_make_aware_dt(date_from),
            created_at__lte=_make_aware_dt(date_to, 23, 59, 59),
        ).count()

    if 'repeat_patients' in metrics:
        # Get distinct patients who had consultations in this date range
        patient_ids = base_qs.values_list('patient', flat=True).distinct()
        # Count their lifetime consultations (all-time, not just in range)
        lifetime_counts = (
            Consultation.objects.filter(patient__in=patient_ids)
            .values('patient')
            .annotate(count=Count('id'))
        )
        repeat_count = sum(1 for p in lifetime_counts if p['count'] > 1)
        new_count = sum(1 for p in lifetime_counts if p['count'] == 1)
        total_unique = repeat_count + new_count
        results['repeat_patients'] = {
            'repeat': repeat_count, 'new': new_count,
            'total': total_unique,
            'repeat_pct': round(repeat_count / total_unique * 100, 1) if total_unique else 0,
        }

    if 'trend' in metrics:
        days = (date_to - date_from).days
        if days <= 31:
            results['trend'] = [
                {
                    'label': day.strftime('%b %d'),
                    'count': base_qs.filter(
                        created_at__gte=_make_aware_dt(day),
                        created_at__lte=_make_aware_dt(day, 23, 59, 59),
                    ).count(),
                }
                for day in [date_from + timedelta(days=i) for i in range(days + 1)]
            ]
        elif days <= 365:
            monthly = (
                base_qs
                .annotate(yr=ExtractYear('created_at'), mo=ExtractMonth('created_at'))
                .values('yr', 'mo')
                .annotate(count=Count('id'))
                .order_by('yr', 'mo')
            )
            results['trend'] = []
            for g in monthly:
                if g['yr'] is not None and g['mo'] is not None:
                    results['trend'].append({
                        'label': date(int(g['yr']), int(g['mo']), 1).strftime('%b %Y'),
                        'count': g['count'],
                    })
        else:
            yearly = (
                base_qs
                .annotate(yr=ExtractYear('created_at'))
                .values('yr')
                .annotate(count=Count('id'))
                .order_by('yr')
            )
            results['trend'] = []
            for g in yearly:
                if g['yr'] is not None:
                    results['trend'].append({
                        'label': str(int(g['yr'])),
                        'count': g['count'],
                    })

    if 'frequent_patients' in metrics:
        results['frequent_patients'] = list(
            base_qs
            .values(
                'patient__first_name', 'patient__last_name',
                'patient__patient_id', 'patient__college__abbreviation',
            )
            .annotate(count=Count('id'))
            .order_by('-count')[:10]
        )

    if 'top_diagnosis_per_college' in metrics:
        per_college = []
        for college in College.objects.all():
            top = (
                Prescription.objects
                .filter(
                    consultation__in=completed_qs,
                    consultation__patient__college=college,
                )
                .values('diagnosis')
                .annotate(count=Count('id'))
                .order_by('-count')
                .first()
            )
            if top:
                per_college.append({
                    'college':   college.abbreviation,
                    'diagnosis': top['diagnosis'],
                    'count':     top['count'],
                })
        results['top_diagnosis_per_college'] = per_college

    if 'low_stock' in metrics:
        results['low_stock'] = list(
            Medicine.objects.filter(quantity__lte=F('low_stock_threshold'))
            .order_by('quantity').values('name', 'quantity', 'low_stock_threshold', 'unit')
        )

    if grouping == 'college':
        results['grouped'] = list(
            base_qs.values('patient__college__abbreviation')
            .annotate(count=Count('id')).order_by('-count')
        )
    elif grouping == 'diagnosis':
        results['grouped'] = list(
            completed_qs.values('prescriptions__diagnosis')
            .annotate(count=Count('id')).order_by('-count')[:20]
        )
    elif grouping == 'date':
        days = (date_to - date_from).days
        if days <= 31:
            # Group by day
            results['grouped'] = [
                {
                    'label': day.strftime('%b %d'),
                    'count': base_qs.filter(
                        created_at__gte=_make_aware_dt(day),
                        created_at__lte=_make_aware_dt(day, 23, 59, 59),
                    ).count(),
                }
                for day in [date_from + timedelta(days=i) for i in range(days + 1)]
            ]
        elif days <= 365:
            # Group by month using Extract
            month_grouped = (
                base_qs
                .annotate(yr=ExtractYear('created_at'), mo=ExtractMonth('created_at'))
                .values('yr', 'mo')
                .annotate(count=Count('id'))
                .order_by('yr', 'mo')
            )
            results['grouped'] = []
            for g in month_grouped:
                if g['yr'] is not None and g['mo'] is not None:
                    results['grouped'].append({
                        'label': date(int(g['yr']), int(g['mo']), 1).strftime('%b %Y'),
                        'count': g['count'],
                    })
        else:
            # Group by year using Extract
            year_grouped = (
                base_qs
                .annotate(yr=ExtractYear('created_at'))
                .values('yr')
                .annotate(count=Count('id'))
                .order_by('yr')
            )
            results['grouped'] = []
            for g in year_grouped:
                if g['yr'] is not None:
                    results['grouped'].append({
                        'label': str(int(g['yr'])),
                        'count': g['count'],
                    })

    return results


def _report_csv(results, date_from, date_to):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="report_{date_from}_{date_to}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Clinic Report', f'{date_from} to {date_to}'])
    writer.writerow([])

    kv_map = [
        ('total_consultations', 'Total Consultations'),
        ('total_patients', 'Total Unique Patients'),
        ('completion_rate', 'Completion Rate (%)'),
        ('cancellation_rate', 'Cancellation Rate (%)'),
        ('avg_per_day', 'Avg Consultations / Day'),
        ('new_patients', 'New Patients in Period'),
    ]
    for key, label in kv_map:
        if key in results:
            writer.writerow([label, results[key]])
    writer.writerow([])

    if 'repeat_patients' in results:
        rp = results['repeat_patients']
        writer.writerow(['Patient Frequency', ''])
        writer.writerow(['New Patients', rp['new']])
        writer.writerow(['Repeat Patients', rp['repeat']])
        writer.writerow(['Repeat %', f"{rp['repeat_pct']}%"])
        writer.writerow([])

    if 'top_diagnoses' in results and results['top_diagnoses']:
        writer.writerow(['Top Diagnoses', ''])
        writer.writerow(['Diagnosis', 'Count'])
        for row in results['top_diagnoses']:
            writer.writerow([row['diagnosis'], row['count']])
        writer.writerow([])

    if 'top_medicines' in results and results['top_medicines']:
        writer.writerow(['Top Medicines', ''])
        writer.writerow(['Medicine', 'Count'])
        for row in results['top_medicines']:
            writer.writerow([row['medicine_name'], row['count']])
        writer.writerow([])

    if 'frequent_patients' in results and results['frequent_patients']:
        writer.writerow(['Most Frequent Patients', ''])
        writer.writerow(['Patient', 'ID', 'College', 'Visits'])
        for row in results['frequent_patients']:
            writer.writerow([
                f"{row['patient__first_name']} {row['patient__last_name']}",
                row['patient__patient_id'] or '—',
                row['patient__college__abbreviation'] or '—',
                row['count'],
            ])
        writer.writerow([])

    if 'top_diagnosis_per_college' in results and results['top_diagnosis_per_college']:
        writer.writerow(['Top Diagnosis per College', ''])
        writer.writerow(['College', 'Diagnosis', 'Count'])
        for row in results['top_diagnosis_per_college']:
            writer.writerow([row['college'], row['diagnosis'], row['count']])
        writer.writerow([])

    if 'trend' in results and results['trend']:
        writer.writerow(['Consultation Trend', ''])
        writer.writerow(['Period', 'Consultations'])
        for row in results['trend']:
            writer.writerow([row.get('label', '—'), row['count']])
        writer.writerow([])

    return response


def _report_excel(results, date_from, date_to):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1D9E75')

    def write_header(ws, row, cols):
        for col_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    current_row = [1]

    def next_row():
        r = current_row[0]
        current_row[0] += 1
        return r

    r = next_row()
    ws.cell(row=r, column=1, value=f'Clinic Report: {date_from} to {date_to}').font = Font(bold=True, size=14)
    next_row()

    def add_kv(label, value):
        r = next_row()
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)

    kv_map = [
        ('total_consultations', 'Total Consultations'),
        ('total_patients', 'Total Unique Patients'),
        ('completion_rate', 'Completion Rate (%)'),
        ('cancellation_rate', 'Cancellation Rate (%)'),
        ('avg_per_day', 'Avg Consultations / Day'),
        ('new_patients', 'New Patients in Period'),
    ]
    for key, label in kv_map:
        if key in results:
            add_kv(label, results[key])

    if 'repeat_patients' in results:
        rp = results['repeat_patients']
        add_kv('New Patients', rp['new'])
        add_kv('Repeat Patients', rp['repeat'])
        add_kv('Repeat %', f"{rp['repeat_pct']}%")

    next_row()

    def add_table(headers, rows_data):
        r = next_row()
        write_header(ws, r, headers)
        for row in rows_data:
            r = next_row()
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=r, column=col_idx, value=val)
        next_row()

    if 'top_diagnoses' in results and results['top_diagnoses']:
        add_table(['Diagnosis', 'Count'], [[r['diagnosis'], r['count']] for r in results['top_diagnoses']])

    if 'top_medicines' in results and results['top_medicines']:
        add_table(['Medicine', 'Count'], [[r['medicine_name'], r['count']] for r in results['top_medicines']])

    if 'frequent_patients' in results and results['frequent_patients']:
        add_table(
            ['Patient', 'ID', 'College', 'Visits'],
            [[
                f"{r['patient__first_name']} {r['patient__last_name']}",
                r['patient__patient_id'] or '—',
                r['patient__college__abbreviation'] or '—',
                r['count'],
            ] for r in results['frequent_patients']]
        )

    if 'top_diagnosis_per_college' in results and results['top_diagnosis_per_college']:
        add_table(
            ['College', 'Diagnosis', 'Count'],
            [[r['college'], r['diagnosis'], r['count']] for r in results['top_diagnosis_per_college']]
        )

    if 'trend' in results and results['trend']:
        add_table(
            ['Period', 'Consultations'],
            [[r.get('label', '—'), r['count']] for r in results['trend']]
        )

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="report_{date_from}_{date_to}.xlsx"'
    return response


def _report_pdf(results, date_from, date_to, user_name=None):
    """Generate a professional PDF for the custom report builder."""
    s = _pdf_styles()
    buf, doc = _pdf_build_doc()
    footer = _make_pdf_footer(user_name)

    meta = [
        f'<b>Generated Date:</b>  {date.today().strftime("%B %d, %Y")}',
        f'<b>Report Period:</b>  {date_from.strftime("%B %d, %Y")} &mdash; {date_to.strftime("%B %d, %Y")}',
    ]

    story = []
    story.extend(_pdf_header_block('Custom Report Summary', meta))

    # ── Summary metrics ──
    kv_pairs = [
        ('total_consultations', 'Total Consultations'),
        ('total_patients', 'Total Unique Patients'),
        ('completion_rate', 'Completion Rate (%)', '{}%'),
        ('cancellation_rate', 'Cancellation Rate (%)', '{}%'),
        ('avg_per_day', 'Avg Consultations / Day'),
        ('new_patients', 'New Patients in Period'),
    ]

    summary_rows = []
    for entry in kv_pairs:
        key = entry[0]
        label = entry[1]
        fmt = entry[2] if len(entry) > 2 else None
        if key in results:
            val = results[key]
            if fmt:
                val = fmt.format(val)
            summary_rows.append([label, str(val)])

    if summary_rows:
        story.append(Paragraph('Summary', s['section_title']))
        story.append(_pdf_make_table(
            ['Metric', 'Count'],
            summary_rows,
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
            h_align='CENTER',
        ))
        story.append(Spacer(1, 8))

    # ── Repeat patient ratio ──
    if 'repeat_patients' in results:
        rp = results['repeat_patients']
        story.append(Paragraph('Patient Retention', s['section_title']))
        story.append(_pdf_make_table(
            ['Category', 'Count', 'Percentage'],
            [
                ['New Patients (1 visit)', str(rp['new']), f"{100 - rp['repeat_pct']}%"],
                ['Repeat Visitors', str(rp['repeat']), f"{rp['repeat_pct']}%"],
            ],
            col_widths=[10*cm, 3.5*cm, 3.5*cm],
            aligns=['left', 'right', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Top diagnoses ──
    if 'top_diagnoses' in results and results['top_diagnoses']:
        story.append(Paragraph('Top Diagnoses', s['section_title']))
        story.append(_pdf_make_table(
            ['Rank', 'Diagnosis', 'Cases'],
            [[str(i + 1), r['diagnosis'][:80], str(r['count'])]
             for i, r in enumerate(results['top_diagnoses'])],
            col_widths=[1.5*cm, 11*cm, 4.5*cm],
            aligns=['center', 'left', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Top medicines ──
    if 'top_medicines' in results and results['top_medicines']:
        story.append(Paragraph('Most Prescribed Medicines', s['section_title']))
        story.append(_pdf_make_table(
            ['Rank', 'Medicine', 'Prescriptions'],
            [[str(i + 1), r['medicine_name'], str(r['count'])]
             for i, r in enumerate(results['top_medicines'])],
            col_widths=[1.5*cm, 11*cm, 4.5*cm],
            aligns=['center', 'left', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Cases per college ──
    if 'cases_per_college' in results and results['cases_per_college']:
        story.append(Paragraph('Cases per College', s['section_title']))
        story.append(_pdf_make_table(
            ['College', 'Cases'],
            [[f"{r['patient__college__abbreviation']} &mdash; {r['patient__college__name']}",
              str(r['count'])]
             for r in results['cases_per_college']],
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Cases by sex ──
    if 'cases_by_sex' in results:
        story.append(Paragraph('Cases by Sex', s['section_title']))
        sex_rows = []
        for row in results['cases_by_sex']:
            sex = {'M': 'Male', 'F': 'Female'}.get(row['patient__sex'], 'Unknown')
            sex_rows.append([sex, str(row['count'])])
        story.append(_pdf_make_table(
            ['Sex', 'Cases'], sex_rows,
            col_widths=[10*cm, 7*cm],
            aligns=['left', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Cases by patient type ──
    if 'cases_by_patient_type' in results:
        t = results['cases_by_patient_type']
        story.append(Paragraph('Cases by Patient Type', s['section_title']))
        story.append(_pdf_make_table(
            ['Type', 'Cases'],
            [['Students', str(t['students'])],
             ['Staff', str(t['staff'])],
             ['Instructors', str(t['instructors'])],
             ['Other', str(t['other'])]],
            col_widths=[10*cm, 7*cm],
            aligns=['left', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Urgency breakdown ──
    if 'urgency_breakdown' in results and results['urgency_breakdown']:
        story.append(Paragraph('Urgency Breakdown (Triage)', s['section_title']))
        story.append(_pdf_make_table(
            ['Urgency', 'Count'],
            [[r['urgency'].capitalize(), str(r['count'])]
             for r in results['urgency_breakdown']],
            col_widths=[10*cm, 7*cm],
            aligns=['left', 'center'],
        ))
        story.append(Spacer(1, 8))

    # ── Medicine dispensing summary ──
    if 'medicine_dispensed' in results and results['medicine_dispensed']:
        story.append(Paragraph('Medicine Dispensing Summary', s['section_title']))
        story.append(_pdf_make_table(
            ['Medicine', 'Unit', 'Total Dispensed'],
            [[r['medicine__name'], r['medicine__unit'], str(r['total_dispensed'])]
             for r in results['medicine_dispensed']],
            col_widths=[8*cm, 4.5*cm, 4.5*cm],
            aligns=['left', 'right', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Low stock ──
    if 'low_stock' in results and results['low_stock']:
        story.append(Paragraph('Low Stock Medicines', s['section_title']))
        story.append(_pdf_make_table(
            ['Medicine', 'Current Stock', 'Threshold'],
            [[r['name'], str(r['quantity']), str(r['low_stock_threshold'])]
             for r in results['low_stock']],
            col_widths=[8*cm, 4.5*cm, 4.5*cm],
            aligns=['left', 'right', 'right'],
        ))
        story.append(Spacer(1, 8))

    # ── Grouped summary ──
    if results.get('grouped'):
        grouping_label = results['grouping'].capitalize()
        story.append(Paragraph(f'Grouped Summary by {grouping_label}', s['section_title']))
        group_header = grouping_label if grouping_label else 'Group'
        story.append(_pdf_make_table(
            [group_header, 'Consultations'],
            [[(r.get('label') or r.get('patient__college__abbreviation') or
               r.get('prescriptions__diagnosis') or '&mdash;'),
              str(r['count'])]
             for r in results['grouped']],
            col_widths=[12*cm, 5*cm],
            aligns=['left', 'right'],
        ))

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="report_{date_from}_{date_to}.pdf"'
    )
    return response